import json
import logging
import lxml
import openpyxl
import os
import re
import shutil
import zipfile
try:
    import zlib
    COMPRESSION = zipfile.ZIP_DEFLATED
except:
    COMPRESSION = zipfile.ZIP_STORED

from  constants import (
    DEF_LOG_LEVEL, DEF_SFX, ENGLISH_COL, FMT_SPEC_STR, JSON_LANG_ROW,
    JSON_LOCALE_FILE_NAME, JSON_ZIP_FILE_NAME, NROWS_CHECK, START_COL,
    START_ROW, XML_ATTR_STR_NAME, XML_CDATA_COL, XML_KEY_COL,
    XML_LANG_FILE_NAME, XML_LANG_ROW, XML_LANG_ENGLISH_CODE, XML_TAG_ROOT,
    XML_TAG_STR, XML_TRANS_COL, XML_ZIP_FILE_NAME
)

ZIPFIle_MODES = {
    zipfile.ZIP_DEFLATED: 'deflated',
    zipfile.ZIP_STORED:   'stored',
}

RE_FMT_SPEC = re.compile( FMT_SPEC_STR )

class _BaseLangTranslate:
    def _is_readable_file(self, path):
        return os.path.isfile( path ) and os.access( path, os.R_OK )

    def _out_json_file_name(self, lang):
        return lang.lower() + '.json'

    def _out_xml_file_name(self, lang):
        return lang.lower(), XML_LANG_FILE_NAME

    def set_log_level(self, level):
        """
        Public method to sey logging level.

        level: string defining level as per the "logging" module. One of
               "DEBUG", "INFO", "WARNING", "ERROR", "CRITICAL"
        """
        self._set_log_level( level )

    def _get_zip_outfile(self, xml=False):
        if not self.filesystem:
            return zipfile.ZipFile(
                XML_ZIP_FILE_NAME if xml else JSON_ZIP_FILE_NAME, mode='w'
            )

    def _read_locale_data(self):
        """
        Read locale data: codes, and names
        """
        if not self._is_readable_file( JSON_LOCALE_FILE_NAME ):
            msg = '"{} is not a readable file'.format(
                JSON_LOCALE_FILE_NAME
            )
            logging.error( msg )
            raise ValueError( msg )

        with open( JSON_LOCALE_FILE_NAME, 'r' ) as finp:
            try:
                vals = json.loads( finp.read() )

                locale_codes = [d['code'] for d in vals]
                locale_names = [d['name'] for d in vals]
                return locale_codes, locale_names
            except ValueError:
                raise

    def _get_locale_name(self, lang, locale_codes, locale_names):
        """
        Returns locale_name corresponding to language

        lang: two-character language name
        locale_codes: locale codes from "locale.json"
        locale_names: locale names from "locale.json". Matchs one-to-one with
               locale_codes
        """
        if lang:
            try:
                idx = locale_codes.index( lang )
                return locale_names[idx]
            except ValueError as e:
                msg - 'Unable to find "{}" in locale code, or an issue in '
                'finding the locale name in locale file "{}"'.format(
                    lang, JSON_LOCALE_FILE_NAME
                )
        else:
            msg = 'Missing language name at col. "{} ({})", row "{}"'.format(
                openpyxl.utils.cell.get_column_letter( column ), column,
                self.json_lang_row
            )
            raise ValueError( msg )

    def _write_json_out_file(
            self, data, path, zoutp, lang, irow=None, column=None
    ):
        """
        Writes the iOS JSON output file

        path: path to output file. If zoutp is not None, this is added to the
              .zip file, and thn deleted
        zoutp: either None, or a zipfile.ZipFile object. If None, the file is
               written directly to the file system
        lang: Language name
        irow: Row number. Can be None, in which case it is not used in info
              message
        column: numeric index of column. Can be None, in which case it is not
              used in info message
        """
        with open( path, 'wb' ) as foutp:
            foutp.write(
                json.dumps( data, indent=4, ensure_ascii=False ).encode(
                    'utf-8'
                )
            )

        col_letter = '' if column is None else \
            openpyxl.utils.cell.get_column_letter( column )
        logging.info(
            'Wrote {} strings in col. {} to JSON for "{}" for language '
            '"{}"'.format( irow or '', col_letter, path, lang )
        )

        if zoutp is not None:
            zoutp.write( path )

            try:
                os.unlink( path )
            except OSError as e:
                logger.warn(
                    'Error in deleting JSON file, "{}", after adding it to '
                    ''.format( path, zoutp.filename )
                )

class AppLangTranslate(_BaseLangTranslate):
    suffix = DEF_SFX

    def _is_writable_dir(self, path):
        return os.path.isdir( path ) and os.access( path, os.W_OK )

    def __init__(
            self, path, start_col=START_COL, end_col=0, start_row=START_ROW,
            end_row=0, json_lang_row=JSON_LANG_ROW, xml_lang_row=XML_LANG_ROW,
            english_col=ENGLISH_COL, xml_cdata_col=XML_CDATA_COL,
            xml_key_col=XML_KEY_COL, xml_trans_col=XML_TRANS_COL,
            stop_on_null=True, stop_on_err=False, filesystem=False
    ):
        """
        path: .xlsx file path. Input file in HelpinOut format
        start_col: starting column
        end_col: Ending column. Zero means last column
        start_row: starting row
        end_row: Ending row. Zero means last row
        stop_on_bn_null: if True, output stops at first cellentry that is None.
             This is needef because the HelpinOut Excel has blankrows at the
             end.
        stop_on_err: if True,processing stops if there is an error in any col.
        filesystem: if True, individual output files are written directly to
             the filesystem, else they are written to a .zip file
        """
        if not self._is_readable_file( path ):
            msg = '"{} is not a readable file'.format( path )
            logging.error( msg )
            raise ValueError( msg )

        self.path = path

        self.start_col = start_col
        self.end_col = end_col

        self.start_row = start_row
        self.end_row = end_row

        self.json_lang_row = json_lang_row
        self.xml_lang_row = xml_lang_row

        self.english_col = english_col

        self.xml_key_col = xml_key_col
        self.xml_cdata_col = xml_cdata_col
        self.xml_trans_col = xml_trans_col

        self.stop_on_null = stop_on_null
        self.stop_on_err = stop_on_err

        self.filesystem = filesystem

        self._set_log_level( DEF_LOG_LEVEL  )

        msg = 'Reading from: "{}". Settings are:\n'
        '\tCols={}.{}'
        '\tRows={},{},\n'
        '\tJSON lang. rows={}\n'
        '\tXML lang. rows={}\n'.format(
            self.path, self.start_col, self.end_col, self.start_row,
            self.end_row, self.json_lang_row, self.xml_lang_row
        )
        logging.info( msg )

    def _set_log_level(self, level):
        """
        sets the log levelin the configuration for the "logging" module.
        """
        val = getattr( logging, level.upper(), None )
        if val is None:
            raise ValueError( 'Invalid log level "{}"'.format( level ) )

        logging.info( 'Setting log. level to: "{}" ({})'.format( level, val ) )
        logging.basicConfig( level=val )

    def _cdata(self, txt):
        """
        Wraps text in CDATA tags

        txtL text of entry
        """
        return '<![CDATA[{}]]>'.format( txt.replace( '\n', '<br/>' ) )

    def _col_to_json(self, column, locale_codes, locale_names, zoutp=None):
        """
        Writes translated strings from one column to JSON

        column: numeric index of column
        zoutp: either None, or a zipfile.ZipFile object. If None, the file is
               written directly to the file system
        locale_codes: locale codes from "locale.json"
        locale_names: locale names from "locale.json". Matchs one-to-one with
               locale_codes
        """
        lang = self.ws.cell( column=column, row=self.json_lang_row )

        try:
            locale_name = self._get_locale_name(
                lang.value, locale_codes, locale_names
            )
        except ValueError:
            raise

        data = { 'Locale_Code': locale_name } 

        for i, row in enumerate(
                range( self.start_row, self.ws.max_row + 1 ), 1
        ):
            name = self.ws.cell(
                column=self.xml_key_col, row=row
            ).value or ''
            if self.stop_on_null and not name:
                break

            english =  self.ws.cell( column=self.english_col, row=row )
            if english.value is not None:
                cell = self.ws.cell( column=column, row=row )

                cdata = self.ws.cell(
                    column=self.xml_cdata_col, row=row
                ).value or ''

                data[name.strip()] = re.sub(
                    RE_FMT_SPEC, '', cell.value or english.value
                )


        try:
            path = self._out_json_file_name( lang.value )
        except OSError:
            raise

        self._write_json_out_file(
            data, path, zoutp, lang.value, irow=i, column=column
        )

    def _col_to_xml(self, column, zoutp=None):
        """
        Writes translated strings from one column to XML

        column: numeric index ofcolumn
        zoutp: either None, or a zipfile.ZipFile object. If None, the file is
               written directly to the file system
        """
        cell = self.ws.cell( column=column, row=self.xml_lang_row )
        if not cell.value:
            msg = 'Missing language name at col. "{} ({})", row "{}"'.format(
                openpyxl.utils.cell.get_column_letter( column ), column,
                self.xml_lang_row
            )
            raise ValueError( msg )

        try:
            dir, fname = self._out_xml_file_name( cell.value )

            if not self._is_writable_dir( dir ):
                os.mkdir( dir )

            path = os.path.join( dir, fname )
        except OSError:
            raise

        root = lxml.etree.Element( XML_TAG_ROOT )
        for i, row in enumerate(
                range( self.start_row, self.ws.max_row + 1 ), 1
        ):
            translatable = self.ws.cell(
                column=self.xml_trans_col, row=row
            ).value
            translatable = True if translatable is None else \
                           bool( translatable )

            if not translatable and dir != XML_LANG_ENGLISH_CODE:
                # Non-translatable strings are output only for Englis
                continue

            name = self.ws.cell(
                column=self.xml_key_col, row=row
            ).value or ''
            if self.stop_on_null and not name:
                break

            cell = self.ws.cell( column=column, row=row )
            cdata = self.ws.cell(
                column=self.xml_cdata_col, row=row
            ).value or ''
            if (cdata == 1 or cdata.lower() == 'yes') and not cell.value:
                # Skip CDATA entries altogether if the language translation is
                # missing
                continue

            if translatable:
                child = lxml.etree.SubElement( root, XML_TAG_STR, name=name )
            else:
                # This will be written only for English: we check
                # "translayable" above, and for non-English languages,
                # continue if it is False 
                child = lxml.etree.SubElement(
                    root, XML_TAG_STR, name=name, translatable='False'
                )

            if cdata == 1 or cdata.lower() == 'yes':
                if cell.value:
                    child.text = self._cdata( cell.value )
            else:
                if cell.value:
                   child.text = cell.value
                else:
                    english =  self.ws.cell( column=self.english_col, row=row )
                    child.text = english.value or ''
 
        with open( path, 'wb' ) as foutp:
            foutp.write(
                lxml.etree.tostring(
                    root, pretty_print=True, encoding='utf-8'
                )
            )

        logging.info(
            'Wrote {} strings in col. {} to XML for "{}" for language '
            '"{}"'.format(
                i, openpyxl.utils.cell.get_column_letter( column ), path,
                cell.value
            )
        )

        if zoutp is not None:
            zoutp.write( path )

            shutil.rmtree( dir )

    def _check_limits(self):
        """
        Sanity check for specified rows and columns
        """
        if self.start_col < self.ws.min_column:
            msg = 'Start column "{}" is less than min. col "{}"'.format(
                self.start_col, self.ws.min_column
            )
            raise ValueError( msg )

        if self.end_col == 0:
            self.end_col = self.ws.max_column
        else:
            if self.end_col > self.ws.max_column:
                msg = 'End column "{}" is greater than min. col "{}"'.format(
                    self.end_col, self.ws.max_column
                )
                raise ValueError( msg )

        if self.start_row < self.ws.min_row:
            msg = 'Start row "{}" is less than min. row "{}"'.format(
                self.start_row, self.ws.min_row
            )
            raise ValueError( msg )

        if self.end_row == 0:
            self.end_row = self.ws.max_row
        else:
            if self.end_row > self.ws.max_row:
                msg = 'End row "{}" is greater than min. row "{}"'.format(
                    self.end_row, self.ws.max_row
                )
                raise ValueError( msg )

    def _col_has_data(self, col):
        rng = range( 1, NROWS_CHECK + 1 )
        return bool(
            list(
                filter(
                    None,
                    [self.ws.cell( column=col, row=row ).value for row in rng]
                )
            )
        )

    def to_out(self, xml=True):
        """
        Writes output language files

        xml: if True, XML output is produced. else JSON
        """
        try:
            self.wb = openpyxl.load_workbook( self.path )
            self.ws = self.wb.active

            self._check_limits()
        except:
            raise

        if not xml:
            try:
                locale_codes, locale_names = self._read_locale_data()
            except ValueError:
                raise

        zoutp = self._get_zip_outfile()

        for col in range( self.start_col, self.end_col + 1 ):
            if not self._col_has_data( col ):
                logging.info(
                    'Skipping col. "{} ({}" which has no data in first '
                    '{} rows'.format(
                        openpyxl.utils.cell.get_column_letter( col ), col,
                        NROWS_CHECK
                    )
                )
                continue

            try:
                if xml:
                    self._col_to_xml( col, zoutp=zoutp )
                else:
                    self._col_to_json(
                        col, locale_codes, locale_names, zoutp=zoutp
                    )
            except (OSError, ValueError) as e:
                logging.error(
                    'Exception in processing. col {}  {}:{}'.format(
                        openpyxl.utils.cell.get_column_letter( col ),
                        e.__class__.__name__, e
                    )
                )
                if self.stop_on_err:
                    if not self.filesystem:
                        zoutp.close()
                    raise

        if not self.filesystem:
            zoutp.close()

    to_xml = to_out

    def to_json(self):
        try:
            self.to_out( xml=False )
        except Exception:
            raise

class XML2JSON(_BaseLangTranslate):
    """
    Converts Android XML language files (either single files, or a .zip of
    multiple XML files) to the corresponding JSON format for iOS.
    """
    def __init__(self, files, stop_on_err=False, filesystem=False):
        """
        files: list of input files. Each is either a path to an Android XML
               language file, named as per convention:
                    values-mr/strings.xml  # For Marathi: usual format
                    mr.xml                 # Also accept single file like this
                OR
                The path to a .zip file containing directories in the Android
                languages .zip file, e.g., in the hierarchy:
                    values/strings.xml     # For English
                    values-hi/strings.xml  # For Hindi
                    ...
        stop_on_err: if True,processing stops if there is an error in any col.
        filesystem: if True, individual output files are written directly to
             the filesystem, else they are written to a .zip file
        """
        self.files = files
        self.filesystem = filesystem
        self.stop_on_err = stop_on_err

    def _get_lang_from_file(self, fname):
        vals = os.path.splitext( fname )
        lvals = len( vals )
        if lvals == 2:
            if vals[1] != '.xml':
                logger.warn(
                    f'File "{fname}" does not have the expected extension, '
                    f'".xml"'
                )
        else:
            logger.warn( f'File "{fname}" does not have an extension' )

        return vals[0]

    def _get_lang_from_dir(self, dir):
        vals = dir.rsplit( '-', maxsplit=1 )
        if vals[0] != 'values':
            logger.warn(
                f'Directory "{dir}" does not have the expected format, '
                f'"values-<lang>"'
            )

        return vals[-1]

    def _get_lang(self, path):
        vals = path.split( os.sep, maxsplit=1 )
        lvals = len( vals )
        if lvals == 1:
            return self._get_lang_from_file( vals[0] )
        else:
            return self._get_lang_from_dir( vals[0] )

    def _get_text(self, elem):
        txt = elem.text

        if 'CDATA' in txt:
            return txt[9:-2]  # Stripped of CDATA tags

        return txt

    def _proc_xml_file(self, zoutp, locale_codes, locale_names, path=None):
        path = path or self.infile

        lang = self._get_lang( path )

        outname = lang + '.json'

        doc = lxml.etree.parse( path )

        root = doc.getroot()
        if root.tag != 'resources':
            logger.warn(
                f'Root element in XML file "{path}" is "{root_tag}" instead '
                f'of "resources"'
            )

        locale_name = self._get_locale_name( lang, locale_codes, locale_names )
        data = { 'Locale_Code': locale_name } 

        for elem in root.xpath( '//string' ):
            name = elem.attrib['name']
            data[name.strip()] = self._get_text( elem )

        self._write_json_out_file( data, outname, zoutp, lang )

    def _proc_zip_file(self, zoutp, locale_codes, locale_names, path=None):
        path = path or self.infile

        with zipfile.ZipFile( path, 'r' ) as zinp:
            for fname in zinp.namelist():
                with zinp.open( fname ) as finp:
                    self._proc_xml_file(
                        zoutp, locale_codes, locale_names, path=fname
                    )

    def to_json(self):
        """
        Writes output JSON files in iOS language format
        """
        try:
            locale_codes, locale_names = self._read_locale_data()
        except ValueError:
            raise

        zoutp = self._get_zip_outfile()

        for f in self.files:
            try:
                self.infile = f

                if zipfile.is_zipfile( self.infile ):
                    self._proc_zip_file( zoutp, locale_codes, locale_names )
                else:
                    # Assume XML file
                    self._proc_xml_file( zoutp, locale_codes, locale_names )
            except Exception as e:
                if stop_on_err:
                    raise
